from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl import Workbook
import time
from pathlib import Path
from config.settings import Settings
from config.logger import configure_logger
from .exceptions import (BrowserClosedError, DownloadFailed,
                        FormSubmitFailed, InvalidDataFormat, ResultsSaveError)

logger = configure_logger()

class RPAChallengeScraper:
    def __init__(self, settings=Settings()):
        self.settings = settings
        self.playwright = sync_playwright().start()
        self._initialize_browser()
        logger.info("Browser initialized")

    def _initialize_browser(self):
        self.browser = self.playwright.chromium.launch(
            headless=self.settings.HEADLESS,
            args=["--start-maximized"],
            channel="chrome"
        )
        self.context = self.browser.new_context(no_viewport=True)
        self.page = self.context.new_page()
        self.page.set_default_timeout(self.settings.TIMEOUT)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        time.sleep(self.settings.SHUTDOWN_DELAY)
        self.context.close()
        self.browser.close()
        self.playwright.stop()
        logger.info("Browser closed")

    def download_file(self):
        try:
            if self.settings.DOWNLOAD_PATH.exists():
                logger.info("File exists, skipping download")
                return
                
            self.page.goto(self.settings.BASE_URL)
            with self.page.expect_download() as download_info:
                self.page.locator("a[href='./assets/downloadFiles/challenge.xlsx']").click()
            download = download_info.value
            download.save_as(self.settings.DOWNLOAD_PATH)
            logger.info(f"File downloaded to {self.settings.DOWNLOAD_PATH}")
        except Exception as e:
            logger.error(f"Download failed: {e}")
            raise DownloadFailed(f"Download failed: {e}")

    def read_excel_data(self):
        try:
            wb = openpyxl.load_workbook(self.settings.DOWNLOAD_PATH)
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            data = [dict(zip(headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]
            logger.info(f"Data read successfully - {len(data)} records")
            return data[:10] 
        except Exception as e:
            logger.error(f"Excel read error: {e}")
            raise InvalidDataFormat(f"Excel read error: {e}")

    def start_challenge(self):
        try:
            self.page.goto(self.settings.BASE_URL)
            self.page.locator("button:text('Start')").click()
            logger.info("Challenge started")
        except Exception as e:
            logger.error(f"Failed to start challenge: {e}")
            raise FormSubmitFailed(f"Failed to start challenge: {e}")

    def _check_completion(self):
        try:
            return self.page.locator("div.congratulations").is_visible(timeout=1000)
        except:
            return False

    def _fill_row(self, row):
        fields = {
            "labelFirstName": "First Name",
            "labelLastName": "Last Name",
            "labelCompanyName": "Company Name", 
            "labelRole": "Role in Company",
            "labelAddress": "Address",
            "labelEmail": "Email",
            "labelPhone": "Phone Number"
        }
        
        for field_name, field_label in fields.items():
            self.page.locator(f"//input[@ng-reflect-name='{field_name}']").fill(str(row.get(field_label, "")))
            time.sleep(0.1)

    def _save_results(self, results):
        try:
            wb = Workbook()
            ws = wb.active
            if results:
                headers = list(results[0].keys())
                ws.append(headers)
                for row in results:
                    ws.append([row.get(h) for h in headers])
            wb.save(self.settings.RESULTS_PATH)
            logger.info(f"Results saved to {self.settings.RESULTS_PATH}")
        except Exception as e:
            logger.error(f"Failed to save results: {e}")
            raise ResultsSaveError(f"Failed to save results: {e}")

    def fill_form(self, data):
        results = []
        try:
            for idx, row in enumerate(data, 1):
                if self._check_completion():
                    logger.info("Challenge completed, stopping early")
                    break
                    
                result = {"status": "success", "error": None}
                try:
                    self._fill_row(row)
                    self.page.locator("input[value='Submit']").click()
                    time.sleep(0.3)
                    logger.info(f"Submitted round {idx}/10")
                except Exception as e:
                    result["status"] = "failed"
                    result["error"] = str(e)
                    logger.error(f"Error in round {idx}: {e}")
                results.append({**row, **result})
            
            self._save_results(results)
            return results
        except Exception as e:
            logger.error(f"Form submission failed: {e}")
            raise FormSubmitFailed(f"Form submission failed: {e}")

    def run(self):
        try:
            self.download_file()
            data = self.read_excel_data()
            self.start_challenge()
            results = self.fill_form(data)
            
            if not self.settings.HEADLESS:
                self.page.screenshot(path=str(self.settings.RESULTS_DIR / "final.png"))
                logger.info("Final screenshot saved")
            
            return results
        except Exception as e:
            logger.error(f"Execution error: {e}")
            raise